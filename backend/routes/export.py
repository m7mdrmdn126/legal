from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Dict, Any, List, Optional
from datetime import datetime, date
import csv
import json
import io
import tempfile
import os
import zipfile

# Optional imports - will work without these packages
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from dependencies.auth import get_current_user, get_admin_user
from models.user import User
from config.database import db_manager
from config.settings import settings

router = APIRouter(prefix="/export", tags=["Export/Import"])

class ExportManager:
    """Comprehensive data export manager"""
    
    def __init__(self):
        # Register Arabic font for PDF if available
        try:
            # You would need to add an Arabic font file to your project
            # pdfmetrics.registerFont(TTFont('Arabic', 'fonts/NotoSansArabic-Regular.ttf'))
            self.arabic_font_available = False
        except:
            self.arabic_font_available = False
    
    def export_cases(self, format: str, date_from: Optional[date] = None, 
                    date_to: Optional[date] = None, status: Optional[str] = None,
                    case_type: Optional[str] = None) -> Dict[str, Any]:
        """Export cases data in various formats"""
        
        # Build query with filters (simplified to match actual schema)
        base_query = """
            SELECT c.id, c.case_number, c.plaintiff, c.defendant,
                   ct.name as case_type, c.judgment_type, c.previous_judgment_id,
                   c.created_at, c.updated_at
            FROM cases c
            JOIN case_types ct ON c.case_type_id = ct.id
            WHERE 1=1
        """
        
        params = []
        
        if date_from:
            base_query += " AND date(c.created_at) >= ?"
            params.append(date_from.isoformat())
        
        if date_to:
            base_query += " AND date(c.created_at) <= ?"
            params.append(date_to.isoformat())
        
        if status:
            base_query += " AND c.case_status = ?"
            params.append(status)
        
        if case_type:
            base_query += " AND ct.name LIKE ?"
            params.append(f"%{case_type}%")
        
        base_query += " ORDER BY c.created_at DESC"
        
        # Get data
        cases = db_manager.execute_query(base_query, tuple(params))
        
        if format == "csv":
            return self._export_to_csv(cases, "cases")
        elif format == "excel":
            if not EXCEL_AVAILABLE:
                raise HTTPException(status_code=400, detail="Excel export requires openpyxl package")
            return self._export_to_excel(cases, "cases")
        elif format == "pdf":
            if not PDF_AVAILABLE:
                raise HTTPException(status_code=400, detail="PDF export requires reportlab package")
            return self._export_to_pdf(cases, "cases")
        elif format == "json":
            return self._export_to_json(cases, "cases")
        else:
            raise HTTPException(status_code=400, detail="Unsupported export format")
    
    def export_sessions(self, format: str, date_from: Optional[date] = None,
                       date_to: Optional[date] = None) -> Dict[str, Any]:
        """Export sessions data"""
        
        base_query = """
            SELECT cs.id, c.case_number, c.plaintiff, c.defendant,
                   cs.session_date, cs.session_time, cs.court_name,
                   cs.session_type, cs.session_status, cs.session_notes,
                   cs.created_at, u.full_name as created_by_name
            FROM case_sessions cs
            JOIN cases c ON cs.case_id = c.id
            LEFT JOIN users u ON cs.created_by = u.id
            WHERE 1=1
        """
        
        params = []
        
        if date_from:
            base_query += " AND date(cs.session_date) >= ?"
            params.append(date_from.isoformat())
        
        if date_to:
            base_query += " AND date(cs.session_date) <= ?"
            params.append(date_to.isoformat())
        
        base_query += " ORDER BY cs.session_date DESC"
        
        sessions = db_manager.execute_query(base_query, tuple(params))
        
        if format == "csv":
            return self._export_to_csv(sessions, "sessions")
        elif format == "excel":
            if not EXCEL_AVAILABLE:
                raise HTTPException(status_code=400, detail="Excel export requires openpyxl package")
            return self._export_to_excel(sessions, "sessions")
        elif format == "pdf":
            if not PDF_AVAILABLE:
                raise HTTPException(status_code=400, detail="PDF export requires reportlab package")
            return self._export_to_pdf(sessions, "sessions")
        elif format == "json":
            return self._export_to_json(sessions, "sessions")
        else:
            raise HTTPException(status_code=400, detail="Unsupported export format")
    
    def export_reports_summary(self, format: str, date_from: Optional[date] = None,
                              date_to: Optional[date] = None) -> Dict[str, Any]:
        """Export comprehensive reports summary"""
        
        # Get various statistics
        stats = self._get_comprehensive_stats(date_from, date_to)
        
        if format == "pdf":
            if not PDF_AVAILABLE:
                raise HTTPException(status_code=400, detail="PDF export requires reportlab package")
            return self._export_summary_to_pdf(stats)
        elif format == "excel":
            if not EXCEL_AVAILABLE:
                raise HTTPException(status_code=400, detail="Excel export requires openpyxl package")
            return self._export_summary_to_excel(stats)
        else:
            raise HTTPException(status_code=400, detail="Summary export only supports PDF and Excel")
    
    def _export_to_csv(self, data: List[Dict], data_type: str) -> Dict[str, Any]:
        """Export data to CSV format"""
        if not data:
            raise HTTPException(status_code=404, detail="No data found for export")
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        content = output.getvalue()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{data_type}_export_{timestamp}.csv"
        
        return {
            "content": content,
            "filename": filename,
            "content_type": "text/csv",
            "size": len(content.encode('utf-8'))
        }
    
    def _export_to_excel(self, data: List[Dict], data_type: str) -> Dict[str, Any]:
        """Export data to Excel format with styling"""
        if not data:
            raise HTTPException(status_code=404, detail="No data found for export")
        
        wb = Workbook()
        ws = wb.active
        ws.title = data_type.title()
        
        # Headers
        headers = list(data[0].keys())
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, record in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=record.get(header, ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{data_type}_export_{timestamp}.xlsx"
        
        return {
            "content": content,
            "filename": filename,
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "size": len(content)
        }
    
    def _export_to_pdf(self, data: List[Dict], data_type: str) -> Dict[str, Any]:
        """Export data to PDF format"""
        if not data:
            raise HTTPException(status_code=404, detail="No data found for export")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        elements = []
        
        # Title
        title = f"{data_type.title()} Export Report"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))
        
        # Add export info
        export_info = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Total Records: {len(data)}"
        elements.append(Paragraph(export_info, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Prepare table data
        headers = list(data[0].keys())
        table_data = [headers]
        
        for record in data:
            row = []
            for header in headers:
                value = str(record.get(header, ''))
                # Truncate long values for PDF display
                if len(value) > 30:
                    value = value[:27] + "..."
                row.append(value)
            table_data.append(row)
        
        # Create table
        table = Table(table_data)
        
        # Style table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        content = buffer.getvalue()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{data_type}_export_{timestamp}.pdf"
        
        return {
            "content": content,
            "filename": filename,
            "content_type": "application/pdf",
            "size": len(content)
        }
    
    def _export_to_json(self, data: List[Dict], data_type: str) -> Dict[str, Any]:
        """Export data to JSON format"""
        if not data:
            raise HTTPException(status_code=404, detail="No data found for export")
        
        export_data = {
            "export_date": datetime.now().isoformat(),
            "data_type": data_type,
            "total_records": len(data),
            "data": data
        }
        
        content = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{data_type}_export_{timestamp}.json"
        
        return {
            "content": content,
            "filename": filename,
            "content_type": "application/json",
            "size": len(content.encode('utf-8'))
        }
    
    def _get_comprehensive_stats(self, date_from: Optional[date] = None,
                                date_to: Optional[date] = None) -> Dict[str, Any]:
        """Get comprehensive statistics for summary report"""
        
        date_filter = ""
        params = []
        
        if date_from:
            date_filter += " AND date(c.created_at) >= ?"
            params.append(date_from.isoformat())
        
        if date_to:
            date_filter += " AND date(c.created_at) <= ?"
            params.append(date_to.isoformat())
        
        # Total counts
        total_cases = db_manager.execute_query(
            f"SELECT COUNT(*) as count FROM cases c WHERE 1=1 {date_filter}", tuple(params)
        )[0]['count']
        
        # Cases by status
        cases_by_status = db_manager.execute_query(
            f"""SELECT case_status, COUNT(*) as count 
               FROM cases c WHERE 1=1 {date_filter}
               GROUP BY case_status ORDER BY count DESC""", tuple(params)
        )
        
        # Cases by type
        cases_by_type = db_manager.execute_query(
            f"""SELECT ct.name, COUNT(c.id) as count 
               FROM case_types ct 
               LEFT JOIN cases c ON ct.id = c.case_type_id 
               WHERE 1=1 {date_filter.replace('c.created_at', 'c.created_at')}
               GROUP BY ct.name ORDER BY count DESC""", tuple(params)
        )
        
        # Monthly trend
        monthly_trend = db_manager.execute_query(
            f"""SELECT strftime('%Y-%m', c.created_at) as month, COUNT(*) as count
               FROM cases c WHERE 1=1 {date_filter}
               GROUP BY month ORDER BY month DESC LIMIT 12""", tuple(params)
        )
        
        return {
            "date_from": date_from.isoformat() if date_from else None,
            "date_to": date_to.isoformat() if date_to else None,
            "total_cases": total_cases,
            "cases_by_status": cases_by_status,
            "cases_by_type": cases_by_type,
            "monthly_trend": monthly_trend
        }
    
    def _export_summary_to_pdf(self, stats: Dict[str, Any]) -> Dict[str, Any]:
        """Export summary to PDF - requires reportlab"""
        # This would be implemented with reportlab when available
        raise HTTPException(status_code=501, detail="PDF export not implemented without reportlab")
    
    def _export_summary_to_excel(self, stats: Dict[str, Any]) -> Dict[str, Any]:
        """Export summary to Excel - requires openpyxl"""
        # This would be implemented with openpyxl when available
        raise HTTPException(status_code=501, detail="Excel export not implemented without openpyxl")

# Initialize export manager
export_manager = ExportManager()

@router.get("/cases")
async def export_cases(
    format: str = Query(..., description="Export format: csv, excel, pdf, json"),
    date_from: Optional[date] = Query(None, description="Start date filter"),
    date_to: Optional[date] = Query(None, description="End date filter"),
    status: Optional[str] = Query(None, description="Case status filter"),
    case_type: Optional[str] = Query(None, description="Case type filter"),
    current_user: User = Depends(get_current_user)
):
    """Export cases data in various formats"""
    
    result = export_manager.export_cases(format, date_from, date_to, status, case_type)
    
    if format in ["excel", "pdf"]:
        return StreamingResponse(
            io.BytesIO(result["content"]),
            media_type=result["content_type"],
            headers={"Content-Disposition": f"attachment; filename={result['filename']}"}
        )
    else:
        return StreamingResponse(
            io.StringIO(result["content"]),
            media_type=result["content_type"],
            headers={"Content-Disposition": f"attachment; filename={result['filename']}"}
        )

@router.get("/sessions")
async def export_sessions(
    format: str = Query(..., description="Export format: csv, excel, pdf, json"),
    date_from: Optional[date] = Query(None, description="Start date filter"),
    date_to: Optional[date] = Query(None, description="End date filter"),
    current_user: User = Depends(get_current_user)
):
    """Export sessions data in various formats"""
    
    result = export_manager.export_sessions(format, date_from, date_to)
    
    if format in ["excel", "pdf"]:
        return StreamingResponse(
            io.BytesIO(result["content"]),
            media_type=result["content_type"],
            headers={"Content-Disposition": f"attachment; filename={result['filename']}"}
        )
    else:
        return StreamingResponse(
            io.StringIO(result["content"]),
            media_type=result["content_type"],
            headers={"Content-Disposition": f"attachment; filename={result['filename']}"}
        )

@router.get("/summary")
async def export_summary_report(
    format: str = Query(..., description="Export format: pdf, excel"),
    date_from: Optional[date] = Query(None, description="Start date filter"),
    date_to: Optional[date] = Query(None, description="End date filter"),
    current_user: User = Depends(get_current_user)
):
    """Export comprehensive summary report"""
    
    if format not in ["pdf", "excel"]:
        raise HTTPException(status_code=400, detail="Summary reports only support PDF and Excel formats")
    
    result = export_manager.export_reports_summary(format, date_from, date_to)
    
    return StreamingResponse(
        io.BytesIO(result["content"]),
        media_type=result["content_type"],
        headers={"Content-Disposition": f"attachment; filename={result['filename']}"}
    )

@router.get("/formats")
async def get_supported_formats(
    current_user: User = Depends(get_current_user)
):
    """Get list of supported export formats"""
    formats = {
        "csv": {
            "name": "CSV",
            "description": "Comma Separated Values",
            "supports": ["cases", "sessions", "notes"],
            "available": True
        },
        "json": {
            "name": "JSON",
            "description": "JavaScript Object Notation", 
            "supports": ["cases", "sessions", "notes"],
            "available": True
        }
    }
    
    if EXCEL_AVAILABLE:
        formats["excel"] = {
            "name": "Excel",
            "description": "Microsoft Excel Spreadsheet",
            "supports": ["cases", "sessions", "notes", "summary"],
            "available": True
        }
    else:
        formats["excel"] = {
            "name": "Excel",
            "description": "Microsoft Excel Spreadsheet (requires openpyxl package)",
            "supports": ["cases", "sessions", "notes", "summary"],
            "available": False
        }
    
    if PDF_AVAILABLE:
        formats["pdf"] = {
            "name": "PDF",
            "description": "Portable Document Format", 
            "supports": ["cases", "sessions", "summary"],
            "available": True
        }
    else:
        formats["pdf"] = {
            "name": "PDF",
            "description": "Portable Document Format (requires reportlab package)",
            "supports": ["cases", "sessions", "summary"],
            "available": False
        }
    
    return {"formats": formats}
